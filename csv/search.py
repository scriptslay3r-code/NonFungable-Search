from openpyxl import Workbook, worksheet, load_workbook

wb: Workbook = load_workbook(filename="data1.xlsx")
ws: worksheet = wb.active

asset_id = '1099738898474'
end_column = ws.max_column
#end_row = ws.max_row
## make error exception for empty cells
end_row = 147
start_column = 1
start_row = 1
column_index = start_column
row_index = start_row

'''while column_index <= end_column:
    # here we just want first row
    print(ws.cell(1, column_index).value)
    column_index += 1
    '''
while row_index <= end_row:
	price = ws.cell(row = row_index, column = 5)
	price = int(price.value)
	price = str(price)
	id = ws.cell(row = row_index, column = 1)
	id = int(id.value)
	id = str(id)
	if id == asset_id:
		print("Yes")
		print("Row Index:" + str(row_index))
		print("Starting Wax Price Should Be: " + price + " WAX")
	row_index += 1
	#value = str(x.value)
	#print(price)
	#print(id)

	