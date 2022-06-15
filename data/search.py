from openpyxl import Workbook, worksheet, load_workbook


    
def get_price(asset):
	wb: Workbook = load_workbook(filename="data1.xlsx")
	ws: worksheet = wb.active
	asset_id = '1099738898474'
	end_column = ws.max_column
	#end_row = ws.max_row
	## make error exception for empty cells
	end_row = 147
	start_column = 1
	start_row = 1
	column_index = start_column
	row_index = start_row

	while row_index <= end_row:
			priceCell = ws.cell(row = row_index, column = 5)
			priceValue = int(priceCell.value)
			priceString = str(priceValue)
			idCell = ws.cell(row = row_index, column = 1)
			idValue = int(idCell.value)
			idString= str(idValue)
			if idString == asset:
				print("Yes")
				price_index = str(row_index)
				print("Row Index: " + price_index)
				print("Starting Wax Price Should Be: " + priceString + " WAX")
				#return("Starting Wax Price Should Be: " + priceString + " WAX")
			
			row_index += 1
			

#get_price('1099738898474')