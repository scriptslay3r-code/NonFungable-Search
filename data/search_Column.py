from openpyxl import *

filePath = "data.xlsx"
wb = load_workbook(filePath)
ws = wb.active

col_idx, row = search_value_in_col_index(ws, "'1099738898474'")

def search_value_in_column(ws, search_string, column="A"):
    for row in range(1, ws.max_row + 1):
        coordinate = "{}{}".format(column, row)
        if ws[coordinate].value == search_string:
            return column, row
    return column, None


def search_value_in_col_idx(ws, search_string, col_idx=1):
    for row in range(1, ws.max_row + 1):
        if ws[row][col_idx].value == search_string:
            return col_idx, row
    return col_idx, None


def search_value_in_row_index(ws, search_string, row=1):
    for cell in ws[row]:
        if cell.value == search_string:
            return cell.column, row
    return None, row
    
    search_value_in_column()
    search_value_in_col_idx()
    search_value_in_row_index